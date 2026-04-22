from __future__ import annotations

import math
import subprocess
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Sequence

import pandas as pd
from openpyxl import load_workbook


PROJECT_DIR = Path(__file__).resolve().parent
WORKBOOK_PATH = PROJECT_DIR / "RE100_CapeVerde_Santiago_v1.xlsx"
RECALC_SCRIPT = PROJECT_DIR / "recalc_excel_workbook.ps1"
TIMESTAMP_SOURCE = PROJECT_DIR / "Santiago_2023_proxy_hourly_load_calibrated.csv"
SCENARIO_TABLE = PROJECT_DIR / "Santiago_scenario_comparison.csv"
FIGURES_DIR = PROJECT_DIR / "scenario_A_figures"


@dataclass
class WorkbookSnapshot:
    wind_mw: float
    pv_mw: float
    start_storage_mwh: float
    storage_mwh: float
    backup_limit_pct: float
    annual_load_mwh: float
    wind_generation_mwh: float
    pv_generation_mwh: float
    biodiesel_backup_generation_mwh: float
    overproduction_mwh: float
    lcoe_eur_per_kwh: float


def run_recalc() -> None:
    cmd = [
        "powershell",
        "-ExecutionPolicy",
        "Bypass",
        "-File",
        str(RECALC_SCRIPT),
        str(WORKBOOK_PATH),
    ]
    subprocess.run(cmd, cwd=PROJECT_DIR, check=True, capture_output=True, text=True)


def read_snapshot() -> WorkbookSnapshot:
    wb = load_workbook(WORKBOOK_PATH, data_only=True, keep_links=False)
    ws = wb["Input & Output"]
    return WorkbookSnapshot(
        wind_mw=float(ws["C5"].value),
        pv_mw=float(ws["C6"].value),
        start_storage_mwh=float(ws["C10"].value),
        storage_mwh=float(ws["C11"].value),
        backup_limit_pct=float(ws["C12"].value),
        annual_load_mwh=float(ws["K6"].value),
        wind_generation_mwh=float(ws["K10"].value),
        pv_generation_mwh=float(ws["K11"].value),
        biodiesel_backup_generation_mwh=float(ws["K17"].value),
        overproduction_mwh=float(ws["C22"].value),
        lcoe_eur_per_kwh=float(ws["O82"].value),
    )


def verify_scenario_a_and_consistency() -> tuple[WorkbookSnapshot, bool]:
    run_recalc()
    snap_1 = read_snapshot()
    run_recalc()
    snap_2 = read_snapshot()

    expected = (40.0, 100.0, 225.0, 450.0, 5.0)
    actual = (
        snap_1.wind_mw,
        snap_1.pv_mw,
        snap_1.start_storage_mwh,
        snap_1.storage_mwh,
        snap_1.backup_limit_pct,
    )
    if actual != expected:
        raise ValueError(
            f"Workbook is not currently on Scenario A. Expected {expected}, found {actual}."
        )

    tolerance = 1e-6
    consistent = all(
        math.isclose(getattr(snap_1, field), getattr(snap_2, field), rel_tol=0.0, abs_tol=tolerance)
        for field in WorkbookSnapshot.__annotations__.keys()
    )
    return snap_1, consistent


def load_hourly_series() -> pd.DataFrame:
    timestamps = pd.read_csv(TIMESTAMP_SOURCE, usecols=["timestamp"])
    time_index = pd.to_datetime(timestamps["timestamp"], utc=True)
    if len(time_index) != 8760:
        raise ValueError(f"Expected 8760 timestamps, found {len(time_index)}")

    wb = load_workbook(WORKBOOK_PATH, data_only=True, keep_links=False)
    load_ws = wb["Hourly Load"]
    wind_ws = wb["Wind and Solar Output"]
    residual_ws = wb["Residual load and storage"]

    data = pd.DataFrame(
        {
            "timestamp": time_index,
            "load_mw": [load_ws[f"G{row}"].value for row in range(7, 8767)],
            "wind_mw": [wind_ws[f"C{row}"].value for row in range(5, 8765)],
            "pv_mw": [wind_ws[f"D{row}"].value for row in range(5, 8765)],
            "backup_mw": [residual_ws[f"H{row}"].value for row in range(7, 8767)],
            "residual_before_storage_mw": [residual_ws[f"J{row}"].value for row in range(7, 8767)],
            "storage_level_mwh": [residual_ws[f"O{row}"].value for row in range(7, 8767)],
            "overproduction_mw": [residual_ws[f"P{row}"].value for row in range(7, 8767)],
        }
    )
    data["vre_mw"] = data["wind_mw"] + data["pv_mw"]
    data["month"] = data["timestamp"].dt.month
    return data


def _series_to_points(values: Sequence[float], width: int, height: int, padding: int) -> list[tuple[float, float]]:
    ymin = min(values)
    ymax = max(values)
    if math.isclose(ymin, ymax):
        ymax = ymin + 1.0
    xspan = max(len(values) - 1, 1)
    points: list[tuple[float, float]] = []
    for idx, value in enumerate(values):
        x = padding + (width - 2 * padding) * idx / xspan
        y = height - padding - (height - 2 * padding) * ((value - ymin) / (ymax - ymin))
        points.append((x, y))
    return points


def _polyline(points: Sequence[tuple[float, float]]) -> str:
    return " ".join(f"{x:.2f},{y:.2f}" for x, y in points)


def _write_svg(path: Path, title: str, body: str, width: int = 1200, height: int = 720) -> None:
    svg = f"""<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}">
<rect x="0" y="0" width="{width}" height="{height}" fill="#f8fafc"/>
<text x="50" y="48" font-family="Arial, sans-serif" font-size="28" font-weight="700" fill="#0f172a">{title}</text>
{body}
</svg>
"""
    path.write_text(svg, encoding="utf-8")


def build_line_chart(
    path: Path,
    title: str,
    subtitle: str,
    xlabels: Sequence[str],
    series: Sequence[tuple[str, Sequence[float], str]],
    ylabel: str,
) -> None:
    width, height, padding = 1200, 720, 90
    all_values = [value for _, values, _ in series for value in values]
    ymin = min(all_values)
    ymax = max(all_values)
    if math.isclose(ymin, ymax):
        ymax = ymin + 1.0

    chart_width = width - 2 * padding
    chart_height = height - 2 * padding - 40
    top = 100
    left = padding
    bottom = top + chart_height

    parts = [
        f'<text x="{left}" y="80" font-family="Arial, sans-serif" font-size="16" fill="#475569">{subtitle}</text>',
        f'<line x1="{left}" y1="{bottom}" x2="{left + chart_width}" y2="{bottom}" stroke="#94a3b8" stroke-width="2"/>',
        f'<line x1="{left}" y1="{top}" x2="{left}" y2="{bottom}" stroke="#94a3b8" stroke-width="2"/>',
        f'<text x="30" y="{top + chart_height / 2:.0f}" transform="rotate(-90 30 {top + chart_height / 2:.0f})" font-family="Arial, sans-serif" font-size="15" fill="#334155">{ylabel}</text>',
    ]

    for tick in range(6):
        value = ymin + (ymax - ymin) * tick / 5
        y = bottom - chart_height * tick / 5
        parts.append(f'<line x1="{left}" y1="{y:.2f}" x2="{left + chart_width}" y2="{y:.2f}" stroke="#e2e8f0" stroke-width="1"/>')
        parts.append(f'<text x="{left - 12}" y="{y + 5:.2f}" text-anchor="end" font-family="Arial, sans-serif" font-size="13" fill="#475569">{value:.1f}</text>')

    if len(xlabels) > 1:
        step = max(1, len(xlabels) // 12)
        for idx in range(0, len(xlabels), step):
            x = left + chart_width * idx / (len(xlabels) - 1)
            parts.append(f'<line x1="{x:.2f}" y1="{bottom}" x2="{x:.2f}" y2="{bottom + 6}" stroke="#94a3b8" stroke-width="1"/>')
            parts.append(f'<text x="{x:.2f}" y="{bottom + 24}" text-anchor="middle" font-family="Arial, sans-serif" font-size="12" fill="#475569">{xlabels[idx]}</text>')

    legend_x = left + 20
    legend_y = height - 40
    for idx, (name, values, color) in enumerate(series):
        pts = []
        for point_idx, value in enumerate(values):
            x = left + chart_width * point_idx / max(len(values) - 1, 1)
            y = bottom - chart_height * ((value - ymin) / (ymax - ymin))
            pts.append((x, y))
        parts.append(f'<polyline fill="none" stroke="{color}" stroke-width="3" points="{_polyline(pts)}"/>')
        lx = legend_x + idx * 220
        parts.append(f'<line x1="{lx}" y1="{legend_y}" x2="{lx + 28}" y2="{legend_y}" stroke="{color}" stroke-width="4"/>')
        parts.append(f'<text x="{lx + 36}" y="{legend_y + 5}" font-family="Arial, sans-serif" font-size="14" fill="#0f172a">{name}</text>')

    _write_svg(path, title, "\n".join(parts), width=width, height=height)


def build_grouped_bar_chart(
    path: Path,
    title: str,
    subtitle: str,
    categories: Sequence[str],
    series: Sequence[tuple[str, Sequence[float], str]],
    ylabel: str,
) -> None:
    width, height = 1200, 720
    padding = 100
    chart_width = width - 2 * padding
    chart_height = height - 2 * padding - 40
    left = padding
    top = 100
    bottom = top + chart_height
    all_values = [value for _, values, _ in series for value in values]
    ymax = max(all_values) * 1.15 if all_values else 1.0
    if ymax <= 0:
        ymax = 1.0

    parts = [
        f'<text x="{left}" y="80" font-family="Arial, sans-serif" font-size="16" fill="#475569">{subtitle}</text>',
        f'<line x1="{left}" y1="{bottom}" x2="{left + chart_width}" y2="{bottom}" stroke="#94a3b8" stroke-width="2"/>',
        f'<line x1="{left}" y1="{top}" x2="{left}" y2="{bottom}" stroke="#94a3b8" stroke-width="2"/>',
        f'<text x="35" y="{top + chart_height / 2:.0f}" transform="rotate(-90 35 {top + chart_height / 2:.0f})" font-family="Arial, sans-serif" font-size="15" fill="#334155">{ylabel}</text>',
    ]

    for tick in range(6):
        value = ymax * tick / 5
        y = bottom - chart_height * tick / 5
        parts.append(f'<line x1="{left}" y1="{y:.2f}" x2="{left + chart_width}" y2="{y:.2f}" stroke="#e2e8f0" stroke-width="1"/>')
        parts.append(f'<text x="{left - 12}" y="{y + 5:.2f}" text-anchor="end" font-family="Arial, sans-serif" font-size="13" fill="#475569">{value:.1f}</text>')

    group_width = chart_width / max(len(categories), 1)
    bar_width = group_width / (len(series) + 1)
    for cat_idx, category in enumerate(categories):
        group_left = left + group_width * cat_idx
        parts.append(f'<text x="{group_left + group_width / 2:.2f}" y="{bottom + 24}" text-anchor="middle" font-family="Arial, sans-serif" font-size="12" fill="#475569">{category}</text>')
        for series_idx, (name, values, color) in enumerate(series):
            value = values[cat_idx]
            bar_height = chart_height * value / ymax
            x = group_left + bar_width * (series_idx + 0.5)
            y = bottom - bar_height
            parts.append(f'<rect x="{x:.2f}" y="{y:.2f}" width="{bar_width * 0.8:.2f}" height="{bar_height:.2f}" fill="{color}" rx="2"/>')

    legend_x = left + 20
    legend_y = height - 40
    for idx, (name, _, color) in enumerate(series):
        lx = legend_x + idx * 220
        parts.append(f'<rect x="{lx}" y="{legend_y - 10}" width="24" height="12" fill="{color}" rx="2"/>')
        parts.append(f'<text x="{lx + 34}" y="{legend_y}" font-family="Arial, sans-serif" font-size="14" fill="#0f172a">{name}</text>')

    _write_svg(path, title, "\n".join(parts), width=width, height=height)


def build_duration_curve(path: Path, title: str, subtitle: str, values: Sequence[float], color: str, ylabel: str) -> None:
    sorted_values = sorted(values, reverse=True)
    xlabels = [str(i) for i in range(len(sorted_values))]
    build_line_chart(
        path=path,
        title=title,
        subtitle=subtitle,
        xlabels=xlabels,
        series=[("Residual load", sorted_values, color)],
        ylabel=ylabel,
    )


def build_figures(data: pd.DataFrame, snapshot: WorkbookSnapshot) -> list[str]:
    FIGURES_DIR.mkdir(exist_ok=True)
    generated: list[str] = []

    weekly = (
        data.set_index("timestamp")[["load_mw", "wind_mw", "pv_mw", "backup_mw", "storage_level_mwh"]]
        .resample("W")
        .mean()
        .reset_index()
    )
    weekly_labels = [ts.strftime("%b") if ts.day <= 7 else "" for ts in weekly["timestamp"]]

    path = FIGURES_DIR / "01_weekly_load_generation.svg"
    build_line_chart(
        path=path,
        title="Scenario A: Weekly Average Load and Generation",
        subtitle=f"Scenario A in workbook: 40 MW wind, 100 MW PV, 450 MWh storage, 5% backup limit. LCOE-like metric = {snapshot.lcoe_eur_per_kwh:.4f} EUR/kWh.",
        xlabels=weekly_labels,
        series=[
            ("Load", weekly["load_mw"].tolist(), "#0f172a"),
            ("Wind", weekly["wind_mw"].tolist(), "#0ea5e9"),
            ("PV", weekly["pv_mw"].tolist(), "#f59e0b"),
            ("Backup", weekly["backup_mw"].tolist(), "#dc2626"),
        ],
        ylabel="MW",
    )
    generated.append(path.name)

    monthly = data.groupby("month").agg(
        load_gwh=("load_mw", lambda s: s.sum() / 1000.0),
        wind_gwh=("wind_mw", lambda s: s.sum() / 1000.0),
        pv_gwh=("pv_mw", lambda s: s.sum() / 1000.0),
        backup_gwh=("backup_mw", lambda s: s.sum() / 1000.0),
        overproduction_gwh=("overproduction_mw", lambda s: s.sum() / 1000.0),
    )
    month_labels = [pd.Timestamp(year=2023, month=month, day=1).strftime("%b") for month in monthly.index]

    path = FIGURES_DIR / "02_monthly_energy_balance.svg"
    build_grouped_bar_chart(
        path=path,
        title="Scenario A: Monthly Energy Balance",
        subtitle="Grouped monthly totals from the populated RE100 workbook output sheets.",
        categories=month_labels,
        series=[
            ("Load", monthly["load_gwh"].tolist(), "#0f172a"),
            ("Wind", monthly["wind_gwh"].tolist(), "#0ea5e9"),
            ("PV", monthly["pv_gwh"].tolist(), "#f59e0b"),
            ("Backup", monthly["backup_gwh"].tolist(), "#dc2626"),
            ("Overproduction", monthly["overproduction_gwh"].tolist(), "#22c55e"),
        ],
        ylabel="GWh/month",
    )
    generated.append(path.name)

    path = FIGURES_DIR / "03_weekly_storage_level.svg"
    build_line_chart(
        path=path,
        title="Scenario A: Weekly Average Storage Level",
        subtitle="Average stored energy in the workbook's storage-level column across 2023.",
        xlabels=weekly_labels,
        series=[("Storage level", weekly["storage_level_mwh"].tolist(), "#7c3aed")],
        ylabel="MWh",
    )
    generated.append(path.name)

    rolling_backup = data["backup_mw"].rolling(24 * 7).sum().fillna(0)
    end_idx = int(rolling_backup.idxmax())
    start_idx = max(0, end_idx - 24 * 7 + 1)
    week_slice = data.iloc[start_idx : start_idx + 24 * 7].copy()
    week_labels = [ts.strftime("%d %b") if ts.hour == 0 else "" for ts in week_slice["timestamp"]]

    path = FIGURES_DIR / "04_peak_backup_week.svg"
    build_line_chart(
        path=path,
        title="Scenario A: Highest-Backup Week",
        subtitle="Hourly slice for the seven-day window with the largest rolling biodiesel backup need.",
        xlabels=week_labels,
        series=[
            ("Load", week_slice["load_mw"].tolist(), "#0f172a"),
            ("Wind + PV", week_slice["vre_mw"].tolist(), "#16a34a"),
            ("Backup", week_slice["backup_mw"].tolist(), "#dc2626"),
        ],
        ylabel="MW",
    )
    generated.append(path.name)

    path = FIGURES_DIR / "05_residual_load_duration_curve.svg"
    build_duration_curve(
        path=path,
        title="Scenario A: Residual Load Duration Curve",
        subtitle="Sorted hourly residual load before storage and backup from the workbook's residual-load sheet.",
        values=data["residual_before_storage_mw"].tolist(),
        color="#2563eb",
        ylabel="MW",
    )
    generated.append(path.name)

    return generated


def write_summaries(snapshot: WorkbookSnapshot, consistent: bool, data: pd.DataFrame, figures: Sequence[str]) -> None:
    summary = f"""# Class-Ready Santiago RE100 Summary

## Case definition

- Case: Santiago island first-pass RE100 workbook run.
- Load basis: calibrated Santiago proxy derived from the Cabo Verde national Electricity Maps load shape.
- Weather basis: 2023 NASA POWER hourly weather for a representative point near Praia after Renewables.ninja rejected 2023 requests.
- Workbook basis: `RE100_CapeVerde_Santiago_v1.xlsx`, recalculated twice in Excel before this pack was written.

## Scenario A currently loaded in the workbook

- Wind capacity: `{snapshot.wind_mw:.0f} MW`
- PV capacity: `{snapshot.pv_mw:.0f} MW`
- Start storage: `{snapshot.start_storage_mwh:.0f} MWh`
- Max storage: `{snapshot.storage_mwh:.0f} MWh`
- Backup limit: `{snapshot.backup_limit_pct:.1f}%`

## Scenario A headline results

- Annual load served in workbook: `{snapshot.annual_load_mwh / 1000:.3f} GWh`
- Annual wind generation: `{snapshot.wind_generation_mwh / 1000:.3f} GWh`
- Annual PV generation: `{snapshot.pv_generation_mwh / 1000:.3f} GWh`
- Annual biodiesel backup generation: `{snapshot.biodiesel_backup_generation_mwh / 1000:.3f} GWh`
- Biodiesel backup share: `{snapshot.biodiesel_backup_generation_mwh / snapshot.annual_load_mwh * 100:.2f}%`
- Annual overproduction: `{snapshot.overproduction_mwh / 1000:.3f} GWh`
- Workbook cost-per-kWh-used metric (`O82`): `{snapshot.lcoe_eur_per_kwh:.4f} EUR/kWh`

## Interpretation

- Scenario A remains the most presentation-ready first-pass option because it stays under the 5% backup limit while keeping the workbook cost lower than the other feasible tested cases.
- Scenario C is still worth discussing as the lower-backup alternative because its backup share is smaller, but it comes with a slightly higher workbook cost metric.
- The workbook's `K16` cell mirrors biodiesel generation in this template, so it is treated here as the energy handed to backup rather than final unserved load.

## Recalculation check

- Workbook recalculation consistency: `{"passed" if consistent else "failed"}`.
- Check method: force full Excel recalculation twice and compare the key Scenario A output cells.

## Figure pack

{chr(10).join(f"- `scenario_A_figures/{name}`" for name in figures)}

## Main cautions for class presentation

- Santiago load is proxy-based, not metered.
- Santiago wind and solar are representative-point weather inputs, not measured plant output.
- Technology cost inputs still come from the course workbook template and should later be localized.
"""
    (PROJECT_DIR / "class_ready_summary.md").write_text(summary, encoding="utf-8")

    short = f"""# Santiago RE100 Short Summary

- Scenario A is currently loaded and verified in the workbook.
- Configuration: `40 MW` wind, `100 MW` PV, `450 MWh` storage, `5%` biodiesel backup limit.
- Result: `{snapshot.annual_load_mwh / 1000:.3f} GWh` load, `{snapshot.biodiesel_backup_generation_mwh / snapshot.annual_load_mwh * 100:.2f}%` backup share, `{snapshot.overproduction_mwh / 1000:.3f} GWh` overproduction, `{snapshot.lcoe_eur_per_kwh:.4f} EUR/kWh`.
- Best classroom message: Scenario A is the strongest first-pass Santiago case because it meets the 5% backup constraint while keeping the workbook cost metric lower than the other feasible options tested.
- Biggest caveat: Santiago load is still proxy-based and should be replaced later with operator data.
"""
    (PROJECT_DIR / "class_ready_summary_short.md").write_text(short, encoding="utf-8")

    figure_index_lines = [
        "# Scenario A Figures Index",
        "",
        "These figures were generated from the populated `RE100_CapeVerde_Santiago_v1.xlsx` workbook while Scenario A was loaded.",
        "",
        f"- Workbook consistency check before export: `{'passed' if consistent else 'failed'}`.",
        "",
        "- `01_weekly_load_generation.svg`: weekly average load, wind, PV, and backup across the year.",
        "- `02_monthly_energy_balance.svg`: monthly load, wind, PV, backup, and overproduction totals.",
        "- `03_weekly_storage_level.svg`: weekly average storage filling level.",
        "- `04_peak_backup_week.svg`: hourly view of the week with the largest rolling backup requirement.",
        "- `05_residual_load_duration_curve.svg`: sorted hourly residual load before storage and backup.",
    ]
    (PROJECT_DIR / "scenario_A_figures_index.md").write_text("\n".join(figure_index_lines) + "\n", encoding="utf-8")

    next_tests = pd.DataFrame(
        [
            ["A_plus_pv", 40, 110, 450, 5.0, "Test whether a small PV increase cuts backup without a large cost penalty."],
            ["A_plus_storage", 40, 100, 600, 5.0, "Check whether moderate extra storage reduces backup and overproduction together."],
            ["A_tighter_backup", 40, 100, 600, 4.0, "See how close the current best shape can get to a stricter backup target."],
            ["A_more_wind", 45, 95, 450, 5.0, "Probe whether a slightly wind-heavier mix improves shoulder-season balance."],
            ["C_plus_storage", 30, 125, 600, 5.0, "Improve the lower-backup Scenario C with more storage to compare against A."],
            ["hybrid_A_C", 35, 110, 500, 5.0, "Blend A and C to test whether backup can fall further while staying near A's cost."],
        ],
        columns=["test_id", "wind_mw", "pv_mw", "storage_mwh", "backup_limit_pct", "why_test_next"],
    )
    next_tests.to_csv(PROJECT_DIR / "scenario_next_tests.csv", index=False)


def main() -> None:
    snapshot, consistent = verify_scenario_a_and_consistency()
    data = load_hourly_series()
    figures = build_figures(data, snapshot)
    write_summaries(snapshot, consistent, data, figures)


if __name__ == "__main__":
    main()
