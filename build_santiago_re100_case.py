from __future__ import annotations

import shutil
import subprocess
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook


PROJECT_DIR = Path(__file__).resolve().parent
MODEL_DIR = PROJECT_DIR / "Model RE100 and handbook"
ORIGINAL_WORKBOOK = MODEL_DIR / "RE100 Model Version 1.0 2026.xlsx"
TARGET_WORKBOOK = PROJECT_DIR / "RE100_CapeVerde_Santiago_v1.xlsx"
RECALC_SCRIPT = PROJECT_DIR / "recalc_excel_workbook.ps1"

LOAD_FILE = PROJECT_DIR / "Santiago_2023_proxy_hourly_load_calibrated.csv"
WIND_FILE = PROJECT_DIR / "Santiago_2023_hourly_wind.csv"
SOLAR_FILE = PROJECT_DIR / "Santiago_2023_hourly_solar.csv"

SCENARIO_COMPARISON_CSV = PROJECT_DIR / "Santiago_scenario_comparison.csv"
SCENARIO_COMPARISON_XLSX = PROJECT_DIR / "Santiago_scenario_comparison.xlsx"
SCENARIO_SUMMARY_MD = PROJECT_DIR / "Santiago_scenario_summary.md"
STUDENT_NOTE_MD = PROJECT_DIR / "explain_like_im_a_student.md"
METHOD_NOTE_MD = PROJECT_DIR / "Santiago_case_method_note.md"

LOAD_PROXY_LABEL = "Santiago calibrated proxy load from Cabo Verde calibrated national shape"
WEATHER_LABEL = "NASA POWER Hourly API fallback after Renewables.ninja 2023 rejection"
REPRESENTATIVE_POINT = "Praia, Santiago Island, Cabo Verde (14.92, -23.51)"


@dataclass
class ScenarioResult:
    scenario: str
    wind_mw: float
    pv_mw: float
    storage_mwh: float
    start_storage_mwh: float
    backup_limit_pct: float
    annual_load_mwh: float
    wind_generation_mwh: float
    pv_generation_mwh: float
    storage_generation_mwh: float
    storage_pumping_mwh: float
    biodiesel_backup_generation_mwh: float
    backup_share_pct: float
    overproduction_mwh: float
    lcoe_eur_per_kwh: float
    residual_load_1_not_met_mwh: float
    residual_load_2_not_met_mwh: float
    max_biodiesel_capacity_mw: float
    max_pump_mw: float
    max_storage_generation_mw: float
    feasible_against_limit: bool
    obvious_constraint_or_failure: str


def round_to_increment(value: float, increment: float, minimum: float | None = None) -> float:
    rounded = round(value / increment) * increment
    if minimum is not None:
        rounded = max(minimum, rounded)
    return float(rounded)


def load_inputs() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    load_df = pd.read_csv(LOAD_FILE)
    wind_df = pd.read_csv(WIND_FILE)
    solar_df = pd.read_csv(SOLAR_FILE)

    if len(load_df) != 8760:
        raise ValueError(f"Expected 8760 Santiago load rows, found {len(load_df)}")
    if len(wind_df) != 8760:
        raise ValueError(f"Expected 8760 wind rows, found {len(wind_df)}")
    if len(solar_df) != 8760:
        raise ValueError(f"Expected 8760 solar rows, found {len(solar_df)}")

    load_ts = pd.to_datetime(load_df["timestamp"], utc=True)
    wind_ts = pd.to_datetime(wind_df["timestamp"], utc=True)
    solar_ts = pd.to_datetime(solar_df["timestamp"], utc=True)
    if not load_ts.equals(wind_ts) or not load_ts.equals(solar_ts):
        raise ValueError("Load, wind, and solar timestamps are not aligned.")

    return load_df, wind_df, solar_df


def copy_original_workbook() -> None:
    shutil.copy2(ORIGINAL_WORKBOOK, TARGET_WORKBOOK)


def _replace_external_formula_artifacts(workbook_path: Path) -> None:
    wb = load_workbook(workbook_path, data_only=False, keep_links=False)
    residual = wb["Residual load and storage"]

    residual["T3"] = "='Input & Output'!B9"
    residual["C4"] = "='Solid Biomass'!B4"
    residual["D4"] = "='Geothermal Power'!B4"
    residual["E4"] = "='Waste to Energy'!B4"

    # These helper columns are not used by the core outputs we extract.
    # Clearing them avoids stale Barbados external-link formulas.
    for row in range(7, 8767):
        for col in range(22, 26):  # V:Y
            residual.cell(row=row, column=col).value = None

    wb.save(workbook_path)


def populate_workbook_with_inputs(
    workbook_path: Path,
    load_df: pd.DataFrame,
    wind_df: pd.DataFrame,
    solar_df: pd.DataFrame,
) -> None:
    wb = load_workbook(workbook_path, data_only=False, keep_links=False)
    io_ws = wb["Input & Output"]
    load_ws = wb["Hourly Load"]
    weather_ws = wb["Wind and Solar Input"]

    # Demand and mobility assumptions: keep Santiago hourly proxy exactly as input.
    io_ws["C26"] = 0
    io_ws["C30"] = 0

    # Load input: place Santiago hourly proxy into the active load path.
    for row_offset, load_value in enumerate(load_df["load_value"], start=7):
        load_ws[f"F{row_offset}"] = float(load_value)
        load_ws[f"H{row_offset}"] = 0.0

    # Wind input: populate both the generic D-column and the operative U-column
    # so summary cells and active generation formulas are consistent.
    weather_ws["D1"] = "Santiago/Praia representative point"
    weather_ws["U1"] = "Santiago/Praia representative point"
    for row_offset, wind_value in enumerate(wind_df["wind_speed_10m_m_per_s"], start=5):
        weather_ws[f"D{row_offset}"] = float(wind_value)
        weather_ws[f"U{row_offset}"] = float(wind_value)

    # Solar input expected by the workbook is hourly radiation in kW/m^2.
    for row_offset, solar_value in enumerate(solar_df["solar_radiation_kw_per_m2"], start=5):
        weather_ws[f"E{row_offset}"] = float(solar_value)

    # Default technology assumptions for a first renewable-only Santiago run.
    io_ws["C5"] = 1.0
    io_ws["C6"] = 1.0
    io_ws["C7"] = 1e-8
    io_ws["C8"] = 1e-8
    io_ws["C9"] = 1e-8
    io_ws["C10"] = 250.0
    io_ws["C11"] = 500.0
    io_ws["C12"] = 5.0
    io_ws["C66"] = 10

    if "Santiago Case Note" in wb.sheetnames:
        del wb["Santiago Case Note"]
    notes = wb.create_sheet("Santiago Case Note")
    notes["A1"] = "Santiago RE100 first-pass workbook note"
    notes["A2"] = "Load source"
    notes["B2"] = LOAD_PROXY_LABEL
    notes["A3"] = "Weather source"
    notes["B3"] = WEATHER_LABEL
    notes["A4"] = "Representative point"
    notes["B4"] = REPRESENTATIVE_POINT
    notes["A5"] = "Important caution"
    notes["B5"] = "Santiago load remains proxy-based, not measured."

    wb.save(workbook_path)


def recalc_workbook(workbook_path: Path) -> None:
    cmd = [
        "powershell",
        "-ExecutionPolicy",
        "Bypass",
        "-File",
        str(RECALC_SCRIPT),
        str(workbook_path),
    ]
    result = subprocess.run(
        cmd,
        cwd=PROJECT_DIR,
        capture_output=True,
        text=True,
        check=True,
    )
    if result.stdout:
        print(result.stdout.strip())


def read_outputs(workbook_path: Path) -> dict[str, float]:
    wb = load_workbook(workbook_path, data_only=True, keep_links=False)
    ws = wb["Input & Output"]

    def val(cell: str) -> float:
        return float(ws[cell].value or 0.0)

    return {
        "annual_load_mwh": val("K6"),
        "wind_generation_mwh": val("K10"),
        "pv_generation_mwh": val("K11"),
        "renewable_generation_mwh": val("K12"),
        "residual_load_1_not_met_mwh": val("K13"),
        "energy_used_for_pumping_mwh": val("K14"),
        "energy_generated_from_storage_mwh": val("K15"),
        "residual_load_2_not_met_mwh": val("K16"),
        "biodiesel_backup_generation_mwh": val("K17"),
        "overproduction_mwh": val("K19"),
        "max_biodiesel_capacity_mw": val("C23"),
        "max_pump_mw": val("L27"),
        "max_storage_generation_mw": val("K27"),
        "final_storage_mwh": val("M27"),
        "storage_pumping_mwh": val("N27"),
        "storage_generation_mwh": val("O27"),
        "lcoe_eur_per_kwh": val("O82"),
        "lcoe_local_per_kwh": val("P82"),
    }


def set_scenario_inputs(
    workbook_path: Path,
    wind_mw: float,
    pv_mw: float,
    storage_mwh: float,
    backup_limit_pct: float,
) -> None:
    wb = load_workbook(workbook_path, data_only=False, keep_links=False)
    ws = wb["Input & Output"]
    ws["C5"] = float(wind_mw)
    ws["C6"] = float(pv_mw)
    ws["C10"] = float(storage_mwh) / 2.0
    ws["C11"] = float(storage_mwh)
    ws["C12"] = float(backup_limit_pct)

    notes = wb["Santiago Case Note"]
    notes["A6"] = "Scenario currently loaded"
    notes["B6"] = f"Wind={wind_mw} MW, PV={pv_mw} MW, Storage={storage_mwh} MWh, Backup limit={backup_limit_pct}%"
    wb.save(workbook_path)


def run_scenario(
    workbook_path: Path,
    name: str,
    wind_mw: float,
    pv_mw: float,
    storage_mwh: float,
    backup_limit_pct: float,
) -> ScenarioResult:
    set_scenario_inputs(workbook_path, wind_mw, pv_mw, storage_mwh, backup_limit_pct)
    recalc_workbook(workbook_path)
    outputs = read_outputs(workbook_path)

    backup_share_pct = (
        outputs["biodiesel_backup_generation_mwh"] / outputs["annual_load_mwh"] * 100.0
        if outputs["annual_load_mwh"] > 0
        else 0.0
    )

    failures: list[str] = []
    if backup_share_pct > backup_limit_pct + 0.01:
        failures.append("Biodiesel share exceeds scenario backup limit")
    if abs(outputs["residual_load_2_not_met_mwh"] - outputs["biodiesel_backup_generation_mwh"]) > 0.1:
        failures.append("Workbook residual-load-to-backup balance changed and should be checked manually")
    if outputs["lcoe_eur_per_kwh"] <= 0:
        failures.append("Workbook returned a non-positive cost metric")
    if not failures:
        failures.append("No obvious first-pass failure")

    return ScenarioResult(
        scenario=name,
        wind_mw=float(wind_mw),
        pv_mw=float(pv_mw),
        storage_mwh=float(storage_mwh),
        start_storage_mwh=float(storage_mwh) / 2.0,
        backup_limit_pct=float(backup_limit_pct),
        annual_load_mwh=outputs["annual_load_mwh"],
        wind_generation_mwh=outputs["wind_generation_mwh"],
        pv_generation_mwh=outputs["pv_generation_mwh"],
        storage_generation_mwh=outputs["storage_generation_mwh"],
        storage_pumping_mwh=outputs["storage_pumping_mwh"],
        biodiesel_backup_generation_mwh=outputs["biodiesel_backup_generation_mwh"],
        backup_share_pct=backup_share_pct,
        overproduction_mwh=outputs["overproduction_mwh"],
        lcoe_eur_per_kwh=outputs["lcoe_eur_per_kwh"],
        residual_load_1_not_met_mwh=outputs["residual_load_1_not_met_mwh"],
        residual_load_2_not_met_mwh=outputs["residual_load_2_not_met_mwh"],
        max_biodiesel_capacity_mw=outputs["max_biodiesel_capacity_mw"],
        max_pump_mw=outputs["max_pump_mw"],
        max_storage_generation_mw=outputs["max_storage_generation_mw"],
        feasible_against_limit=(backup_share_pct <= backup_limit_pct + 0.01),
        obvious_constraint_or_failure="; ".join(failures),
    )


def choose_best_result(results: Iterable[ScenarioResult]) -> ScenarioResult:
    def rank_key(result: ScenarioResult) -> tuple[float, float, float, float]:
        backup_penalty = max(result.backup_share_pct - result.backup_limit_pct, 0.0)
        feasibility_penalty = 0.0 if result.feasible_against_limit else 1.0
        return (
            feasibility_penalty,
            backup_penalty,
            result.lcoe_eur_per_kwh,
            result.overproduction_mwh,
        )

    return sorted(results, key=rank_key)[0]


def build_initial_scenarios(
    annual_load_mwh: float,
    wind_flh: float,
    pv_flh: float,
) -> tuple[float, list[tuple[str, float, float, float, float]]]:
    avg_load_mw = annual_load_mwh / 8760.0
    moderate_storage_mwh = round_to_increment(avg_load_mw * 12.0, 50.0, minimum=100.0)

    factor_candidates = [1.15, 1.25, 1.35]
    shares = {
        "A": (0.50, 0.50),
        "B": (0.65, 0.35),
        "C": (0.35, 0.65),
    }

    scenarios_by_factor: dict[float, list[tuple[str, float, float, float, float]]] = {}
    for factor in factor_candidates:
        scenario_defs = []
        target_generation_mwh = annual_load_mwh * factor
        for name, (wind_share, pv_share) in shares.items():
            wind_mw = round_to_increment(target_generation_mwh * wind_share / wind_flh, 5.0, minimum=5.0)
            pv_mw = round_to_increment(target_generation_mwh * pv_share / pv_flh, 5.0, minimum=5.0)
            scenario_defs.append((name, wind_mw, pv_mw, moderate_storage_mwh, 5.0))
        scenarios_by_factor[factor] = scenario_defs

    return moderate_storage_mwh, scenarios_by_factor


def scenario_results_to_frame(results: list[ScenarioResult]) -> pd.DataFrame:
    frame = pd.DataFrame([asdict(result) for result in results])
    preferred_columns = [
        "scenario",
        "wind_mw",
        "pv_mw",
        "storage_mwh",
        "start_storage_mwh",
        "backup_limit_pct",
        "annual_load_mwh",
        "wind_generation_mwh",
        "pv_generation_mwh",
        "storage_generation_mwh",
        "storage_pumping_mwh",
        "biodiesel_backup_generation_mwh",
        "backup_share_pct",
        "overproduction_mwh",
        "lcoe_eur_per_kwh",
        "residual_load_1_not_met_mwh",
        "residual_load_2_not_met_mwh",
        "max_biodiesel_capacity_mw",
        "max_pump_mw",
        "max_storage_generation_mw",
        "feasible_against_limit",
        "obvious_constraint_or_failure",
    ]
    return frame[preferred_columns]


def write_scenario_outputs(
    scenarios_df: pd.DataFrame,
    best_ac: ScenarioResult,
    best_overall: ScenarioResult,
    wind_flh: float,
    pv_flh: float,
    moderate_storage_mwh: float,
) -> None:
    scenarios_df.to_csv(SCENARIO_COMPARISON_CSV, index=False)

    metadata = pd.DataFrame(
        [
            ["Load source", LOAD_PROXY_LABEL],
            ["Weather source", WEATHER_LABEL],
            ["Representative point", REPRESENTATIVE_POINT],
            ["Wind full-load-equivalent hours from workbook probe", wind_flh],
            ["PV full-load-equivalent hours from workbook probe", pv_flh],
            ["Moderate storage rule", f"{moderate_storage_mwh} MWh (~12 hours of average Santiago load)"],
            ["Best of A-C", best_ac.scenario],
            ["Best overall first-pass scenario", best_overall.scenario],
        ],
        columns=["item", "value"],
    )

    with pd.ExcelWriter(SCENARIO_COMPARISON_XLSX, engine="openpyxl") as writer:
        scenarios_df.to_excel(writer, index=False, sheet_name="Scenarios")
        metadata.to_excel(writer, index=False, sheet_name="Metadata")


def write_method_note(best_overall: ScenarioResult, load_total_mwh: float) -> None:
    text = f"""# Santiago Case Method Note

- Workbook used: `RE100 Model Version 1.0 2026.xlsx`, copied to `RE100_CapeVerde_Santiago_v1.xlsx`.
- Target case: Santiago island.
- Santiago load input: `{LOAD_PROXY_LABEL}`.
- Santiago load status: proxy, not measured.
- Santiago calibrated annual load used in workbook: `{load_total_mwh / 1000:.3f} GWh`.
- Weather source requested first: `Renewables.ninja`.
- Renewables.ninja outcome: rejected 2023 hourly request because the live API only allowed dates through `2019-12-31`.
- Weather source actually used: `{WEATHER_LABEL}`.
- Wind series inserted: hourly `WS10M` at 10 m in UTC.
- Solar series inserted: hourly `ALLSKY_SFC_SW_DWN`, converted from Wh/m^2 to kW/m^2 for workbook compatibility.
- Representative location used for Santiago weather: `{REPRESENTATIVE_POINT}`.
- Storage start-volume assumption for every scenario: `50%` of maximum storage volume.
- Biomass, geothermal, and waste-to-energy capacities were kept effectively at zero for this first-pass scenario batch.
- Workbook hygiene step: stale Barbados external-link artifacts were cleared or replaced before recalculation.
- Workbook interpretation note: in this template, `Residual Load 2 not met` (`K16`) numerically matches biodiesel generation (`K17`), so it is treated here as the amount handed off to backup rather than final unserved energy.
- Workbook result currently loaded in `RE100_CapeVerde_Santiago_v1.xlsx`: scenario `{best_overall.scenario}` with wind `{best_overall.wind_mw:.0f} MW`, PV `{best_overall.pv_mw:.0f} MW`, storage `{best_overall.storage_mwh:.0f} MWh`, backup limit `{best_overall.backup_limit_pct:.1f}%`.
"""
    METHOD_NOTE_MD.write_text(text, encoding="utf-8")


def write_summary_markdown(scenarios: list[ScenarioResult], best_ac: ScenarioResult, best_overall: ScenarioResult) -> None:
    lines = [
        "# Santiago Scenario Summary",
        "",
        f"- Most promising first-pass scenario: `{best_overall.scenario}`.",
        f"- Best scenario among A-C before storage sensitivity: `{best_ac.scenario}`.",
        "- Interpretation rule used: prioritize scenarios that respect the stated backup limit and then minimize workbook cost per kWh used; in this template `K16` is interpreted as the deficit passed to backup rather than final unserved load.",
        "",
        "## Scenario results",
        "",
    ]

    for result in scenarios:
        lines.extend(
            [
                f"### Scenario {result.scenario}",
                "",
                f"- Wind: `{result.wind_mw:.0f} MW`",
                f"- PV: `{result.pv_mw:.0f} MW`",
                f"- Storage: `{result.storage_mwh:.0f} MWh`",
                f"- Backup limit: `{result.backup_limit_pct:.1f}%`",
                f"- Biodiesel backup generation: `{result.biodiesel_backup_generation_mwh / 1000:.3f} GWh`",
                f"- Backup share: `{result.backup_share_pct:.2f}%`",
                f"- Overproduction: `{result.overproduction_mwh / 1000:.3f} GWh`",
                f"- LCOE-like workbook metric (`O82`): `{result.lcoe_eur_per_kwh:.4f} EUR/kWh`",
                f"- Remaining deficit handed to backup in workbook cell `K16`: `{result.residual_load_2_not_met_mwh:.3f} MWh`",
                f"- First-pass assessment: `{result.obvious_constraint_or_failure}`",
                "",
            ]
        )

    SCENARIO_SUMMARY_MD.write_text("\n".join(lines), encoding="utf-8")


def write_student_note(best_overall: ScenarioResult) -> None:
    text = f"""# Explain Like I'm A Student

This workbook run uses three kinds of inputs.

- The load curve is based on a real Cabo Verde hourly series from Electricity Maps, but Santiago itself did not exist as a separate zone, so the Santiago load is a modeled proxy based on the calibrated national shape.
- The wind and solar inputs are not plant meter data from Santiago. Renewables.ninja was tested first, but it would not supply 2023 hourly data. So the weather inputs came from NASA POWER for a representative point near Praia on Santiago.
- The scenario results themselves come from the RE100 workbook logic after the Santiago inputs were inserted and the workbook was recalculated in Excel.

What was real:

- The underlying Cabo Verde national hourly load shape.
- The calibrated national annual electricity scale.
- The 2023 NASA POWER hourly weather record used for the Santiago representative point.

What was proxy or modeled:

- Santiago hourly electricity load.
- Santiago weather representativeness, because one point near Praia stands in for the whole island.
- The scenario capacities themselves, which are planning assumptions rather than historical facts.

What the first scenario results mean:

- Each scenario tests a different mix of wind, PV, and storage for meeting the Santiago proxy load.
- Biodiesel is treated as backup and its share is compared against the scenario limit.
- Storage helps shift excess wind/PV energy into later hours, so the key trade-off is between backup share, overproduction, and cost.
- The workbook's cost result is the `Cost per kWh used` metric from cell `O82`, which works like a workbook-native LCOE-style indicator.

What still needs improvement later:

- Real operator load data for Santiago instead of a population-based proxy.
- Real island weather or plant production data instead of a single representative reanalysis point.
- Local Santiago technology-cost assumptions instead of the template defaults inherited from the course workbook.
- A broader scenario search after the first-pass batch to check whether the current "best" case remains best under more combinations.

At this stage, scenario `{best_overall.scenario}` looks most promising in the first pass because it gives the best balance of backup compliance, residual-load coverage, and workbook cost among the tested cases.
"""
    STUDENT_NOTE_MD.write_text(text, encoding="utf-8")


def main() -> None:
    load_df, wind_df, solar_df = load_inputs()

    copy_original_workbook()
    _replace_external_formula_artifacts(TARGET_WORKBOOK)
    populate_workbook_with_inputs(TARGET_WORKBOOK, load_df, wind_df, solar_df)
    recalc_workbook(TARGET_WORKBOOK)

    probe_outputs = run_scenario(
        TARGET_WORKBOOK,
        name="probe",
        wind_mw=1.0,
        pv_mw=1.0,
        storage_mwh=100.0,
        backup_limit_pct=100.0,
    )

    wind_flh = probe_outputs.wind_generation_mwh / probe_outputs.wind_mw
    pv_flh = probe_outputs.pv_generation_mwh / probe_outputs.pv_mw

    moderate_storage_mwh, scenario_candidates = build_initial_scenarios(
        annual_load_mwh=probe_outputs.annual_load_mwh,
        wind_flh=wind_flh,
        pv_flh=pv_flh,
    )

    ac_results: list[ScenarioResult] = []
    chosen_factor_results: list[ScenarioResult] | None = None
    for factor, definitions in scenario_candidates.items():
        current_results = [
            run_scenario(TARGET_WORKBOOK, name, wind_mw, pv_mw, storage_mwh, backup_limit_pct)
            for name, wind_mw, pv_mw, storage_mwh, backup_limit_pct in definitions
        ]
        feasible = [result for result in current_results if result.feasible_against_limit]
        chosen_factor_results = current_results
        if feasible:
            ac_results = current_results
            break
    if not ac_results:
        ac_results = chosen_factor_results or []

    best_ac = choose_best_result(ac_results)
    larger_storage_mwh = round_to_increment(best_ac.storage_mwh * 1.5, 50.0, minimum=100.0)
    smaller_storage_mwh = round_to_increment(best_ac.storage_mwh * 0.6, 50.0, minimum=50.0)

    scenario_d = run_scenario(
        TARGET_WORKBOOK,
        "D",
        best_ac.wind_mw,
        best_ac.pv_mw,
        larger_storage_mwh,
        2.5,
    )
    scenario_e = run_scenario(
        TARGET_WORKBOOK,
        "E",
        best_ac.wind_mw,
        best_ac.pv_mw,
        smaller_storage_mwh,
        5.0,
    )

    all_results = ac_results + [scenario_d, scenario_e]
    best_overall = choose_best_result(all_results)

    # Save the workbook with the best first-pass scenario loaded.
    set_scenario_inputs(
        TARGET_WORKBOOK,
        best_overall.wind_mw,
        best_overall.pv_mw,
        best_overall.storage_mwh,
        best_overall.backup_limit_pct,
    )
    recalc_workbook(TARGET_WORKBOOK)

    scenarios_df = scenario_results_to_frame(all_results)
    write_scenario_outputs(
        scenarios_df=scenarios_df,
        best_ac=best_ac,
        best_overall=best_overall,
        wind_flh=wind_flh,
        pv_flh=pv_flh,
        moderate_storage_mwh=moderate_storage_mwh,
    )
    write_method_note(best_overall=best_overall, load_total_mwh=probe_outputs.annual_load_mwh)
    write_summary_markdown(all_results, best_ac, best_overall)
    write_student_note(best_overall)


if __name__ == "__main__":
    main()
